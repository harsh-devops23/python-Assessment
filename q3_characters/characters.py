"""
Q3 - Characters of Ice and Fire
=================================
Tasks:
    a. Get all characters from the API
    b. Find how many seasons (tvSeries) each character appeared in
    c. Sort characters by number of season appearances (descending)
    d. Write all sorted data to an Excel file

API: https://anapioficeandfire.com/api/characters
Author: Intern Assignment
"""

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ─── Configuration ────────────────────────────────────────────
BASE_URL    = "https://anapioficeandfire.com/api/characters"
OUTPUT_FILE = "characters_output.xlsx"
PAGE_SIZE   = 50


def fetch_all_characters():
    """
    Fetches ALL characters from the API page by page.
    The API paginates results — we loop until no more data is returned.
    Characters with no name are skipped (API has many unnamed entries).
    """
    all_characters = []
    page           = 1

    print("Fetching characters from API (this may take a minute — large dataset)...")

    while True:
        params   = {"page": page, "pageSize": PAGE_SIZE}
        response = requests.get(BASE_URL, params=params)

        if response.status_code != 200:
            print(f"  Error on page {page}: HTTP {response.status_code}")
            break

        data = response.json()

        if not data:
            break

        all_characters.extend(data)
        print(f"  Page {page} fetched — {len(data)} characters | Total so far: {len(all_characters)}")
        page += 1

    print(f"\nTotal characters fetched: {len(all_characters)}")
    return all_characters


def process_characters(raw_characters):
    """
    Processes raw API data into a clean list of character records.

    For each character, we extract:
        - name          → character name (or "Unnamed" if blank)
        - gender        → gender field from API
        - culture       → cultural background
        - born          → birth info
        - died          → death info
        - titles        → list joined as string
        - aliases       → list joined as string
        - tv_seasons    → list of seasons they appeared in (tvSeries field)
        - season_count  → how many seasons they appeared in (length of tv_seasons)
        - books_count   → how many books they appeared in
        - alive         → True if 'died' field is empty

    Only characters with at least 1 season appearance are kept,
    since the question asks to sort by season appearances.
    """
    processed = []

    for char in raw_characters:
        name         = char.get("name", "").strip()
        gender       = char.get("gender", "Unknown")
        culture      = char.get("culture", "Unknown")
        born         = char.get("born", "Unknown")
        died         = char.get("died", "")
        titles       = ", ".join(char.get("titles", [])) or "None"
        aliases      = ", ".join(char.get("aliases", [])) or "None"
        tv_seasons   = char.get("tvSeries", [])
        books        = char.get("books", [])
        povbooks     = char.get("povBooks", [])

        # Clean up empty season strings the API sometimes returns
        tv_seasons   = [s for s in tv_seasons if s.strip()]
        season_count = len(tv_seasons)
        seasons_str  = ", ".join(tv_seasons) if tv_seasons else "None"
        books_count  = len(books) + len(povbooks)
        alive        = "Deceased" if died else "Alive"

        # Use alias as display name if actual name is blank
        display_name = name if name else (aliases.split(",")[0].strip() if aliases != "None" else "Unknown")

        # Only include characters who appeared in at least 1 season
        if season_count > 0:
            processed.append({
                "name"         : display_name,
                "gender"       : gender,
                "culture"      : culture,
                "born"         : born if born else "Unknown",
                "status"       : alive,
                "titles"       : titles,
                "aliases"      : aliases,
                "season_count" : season_count,
                "seasons"      : seasons_str,
                "books_count"  : books_count,
            })

    return processed


def sort_by_season_appearances(characters):
    """
    Sorts the character list by season_count in descending order.
    Characters who appeared in the most seasons come first.
    """
    return sorted(characters, key=lambda c: c["season_count"], reverse=True)


def write_to_excel(sorted_characters, filename):
    """
    Writes all character data to a well-formatted Excel (.xlsx) file.

    Sheet layout:
        - Row 1: Title header (merged cells)
        - Row 2: Column headers (styled in dark color)
        - Row 3+: Data rows (alternating colors for readability)

    Columns:
        Rank | Name | Gender | Culture | Status | Titles |
        Aliases | Season Count | Seasons Appeared | Books Count
    """
    filepath = os.path.join(os.path.dirname(__file__), filename)
    wb       = openpyxl.Workbook()
    ws       = wb.active
    ws.title = "Characters"

    # ── Color palette ──────────────────────────────────────────
    DARK_RED    = "8B0000"   # Header background
    GOLD        = "FFD700"   # Header text
    LIGHT_RED   = "FFCCCC"   # Odd row background
    WHITE       = "FFFFFF"   # Even row background
    BLACK       = "000000"   # Normal text

    # ── Define column headers and widths ───────────────────────
    columns = [
        ("Rank",             6),
        ("Character Name",   28),
        ("Gender",           10),
        ("Culture",          18),
        ("Status",           10),
        ("Titles",           30),
        ("Aliases",          25),
        ("Season Count",     14),
        ("Seasons Appeared", 40),
        ("Books Count",      13),
    ]

    # ── Row 1: Big merged title row ────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell              = ws.cell(row=1, column=1)
    title_cell.value        = "CHARACTERS OF ICE AND FIRE — Sorted by Season Appearances"
    title_cell.font         = Font(name="Arial", bold=True, size=14, color=GOLD)
    title_cell.fill         = PatternFill("solid", fgColor=DARK_RED)
    title_cell.alignment    = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Row 2: Column headers ──────────────────────────────────
    header_font  = Font(name="Arial", bold=True, size=10, color=GOLD)
    header_fill  = PatternFill("solid", fgColor="5C0000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header_text, col_width) in enumerate(columns, start=1):
        cell              = ws.cell(row=2, column=col_idx, value=header_text)
        cell.font         = header_font
        cell.fill         = header_fill
        cell.alignment    = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[2].height = 22

    # ── Row 3+: Data rows ──────────────────────────────────────
    thin_border = Border(
        left   = Side(style="thin"),
        right  = Side(style="thin"),
        top    = Side(style="thin"),
        bottom = Side(style="thin"),
    )

    for row_idx, char in enumerate(sorted_characters, start=1):
        excel_row    = row_idx + 2
        row_fill     = PatternFill("solid", fgColor=LIGHT_RED if row_idx % 2 != 0 else WHITE)
        data_align   = Alignment(vertical="center", wrap_text=True)

        row_data = [
            row_idx,
            char["name"],
            char["gender"],
            char["culture"],
            char["status"],
            char["titles"],
            char["aliases"],
            char["season_count"],
            char["seasons"],
            char["books_count"],
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = data_align
            cell.font      = Font(name="Arial", size=9)

            # Bold and center the Rank and Season Count columns
            if col_idx in (1, 8):
                cell.font      = Font(name="Arial", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Freeze top 2 rows so headers stay visible when scrolling ──
    ws.freeze_panes = "A3"

    wb.save(filepath)
    print(f"Excel file written to: {filepath}")
    return filepath


def display_preview(sorted_characters, preview_count=15):
    """
    Prints a console preview of the top N characters by season appearances.
    """
    print(f"\n{'='*70}")
    print(f"  TOP {preview_count} CHARACTERS BY SEASON APPEARANCES")
    print(f"{'='*70}")
    print(f"{'Rank':<6} {'Name':<28} {'Seasons':<10} {'Status'}")
    print(f"{'-'*70}")

    for i, char in enumerate(sorted_characters[:preview_count], start=1):
        print(f"{i:<6} {char['name']:<28} {char['season_count']:<10} {char['status']}")

    print(f"{'-'*70}")
    print(f"  Full data for all {len(sorted_characters)} characters saved to Excel.\n")


def main():
    print("=" * 70)
    print("  Q3 — Characters of Ice and Fire")
    print("=" * 70)

    # Step a — Fetch all characters
    raw_characters = fetch_all_characters()

    # Step b — Process and count season appearances
    characters = process_characters(raw_characters)
    print(f"\nCharacters with TV season data: {len(characters)}")

    # Step c — Sort by season appearances (most to least)
    sorted_characters = sort_by_season_appearances(characters)
    print(f"Sorted by season appearances (descending).")

    # Step d — Write to Excel
    write_to_excel(sorted_characters, OUTPUT_FILE)

    # Console preview
    display_preview(sorted_characters)

    print("Q3 Complete! Check 'characters_output.xlsx' for the full Excel file.")


if __name__ == "__main__":
    main()
